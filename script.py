import os
from datetime import datetime
from dotenv import load_dotenv
import requests
from openpyxl import load_workbook
import paramiko

# Load environment variables
load_dotenv()

def get_api_credentials():
    return {
        'client_key': os.getenv('CLIENT_KEY'),
        'client_secret': os.getenv('CLIENT_SECRET'),
        'account_name': os.getenv('ACCOUNT_NAME'),
        'account_id': os.getenv('ACCOUNT_ID')
    }

def get_sftp_config():
    return {
        'hostname': os.getenv('SFTP_HOSTNAME'),
        'port': int(os.getenv('SFTP_PORT', 22)),
        'username': os.getenv('SFTP_USERNAME'),
        'password': os.getenv('SFTP_PASSWORD'),
        'remote_path': os.getenv('SFTP_REMOTE_PATH', '/')
    }

def fetch_api_data(api_url, credentials):
    """Fetch data from API using basic auth"""
    try:
        response = requests.get(
            api_url,
            auth=(credentials['client_key'], credentials['client_secret'])
        )
        response.raise_for_status()
        return response.json()
    except requests.exceptions.RequestException as e:
        print(f"Error fetching API data: {e}")
        return None

def format_date(date_str):
    """Convert ISO date string to mm/dd/yyyy format"""
    if not date_str:
        return ""
    try:
        date_obj = datetime.strptime(date_str.split('T')[0], '%Y-%m-%d')
        return date_obj.strftime('%m/%d/%Y')
    except (ValueError, AttributeError):
        return ""

def minutes_to_hours(minutes):
    """Convert minutes to hours with 2 decimal places"""
    if not minutes:
        return 0
    return round(float(minutes) / 60, 2)

def populate_excel(data, excel_file):
    """Populate Excel file with API data"""
    # Load the workbook
    wb = load_workbook(excel_file)
    ws = wb.active

    # Start from row 6 (after headers)
    row = 6

    # Process each result
    for item in data['results']:
        # Column B: user_email
        ws[f'B{row}'] = item['user_email']

        # Column C: Current date in mm/dd/yyyy format
        ws[f'C{row}'] = datetime.now().strftime('%m/%d/%Y')

        # Column E: course_category
        ws[f'E{row}'] = item['course_category']

        # Column F: course_category (same as E)
        ws[f'F{row}'] = item['course_category']

        # Column G: course_title
        ws[f'G{row}'] = item['course_title']

        # Column H: course_enroll_date
        ws[f'H{row}'] = format_date(item['course_enroll_date'])

        # Column I: course_first_completion_date
        ws[f'I{row}'] = format_date(item['course_first_completion_date'])

        # Column J: Static value "Online_Complete"
        ws[f'J{row}'] = "Online_Complete"

        # Column L: num_video_consumed_minutes converted to hours
        ws[f'L{row}'] = minutes_to_hours(item['num_video_consumed_minutes'])

        row += 1

    # Save the workbook with a new name
    output_file = f"populated_{os.path.basename(excel_file)}"
    wb.save(output_file)
    print(f"Excel file has been populated and saved as: {output_file}")
    return output_file

def upload_to_sftp(local_file, sftp_config):
    """Upload file to SFTP server"""
    try:
        transport = paramiko.Transport((sftp_config['hostname'], sftp_config['port']))
        transport.connect(username=sftp_config['username'], password=sftp_config['password'])

        sftp = paramiko.SFTPClient.from_transport(transport)

        remote_path = os.path.join(sftp_config['remote_path'], os.path.basename(local_file))
        sftp.put(local_file, remote_path)

        print(f"File uploaded successfully to SFTP: {remote_path}")

        sftp.close()
        transport.close()
    except Exception as e:
        print(f"Error uploading file to SFTP: {e}")

def main():
    # Get API credentials
    credentials = get_api_credentials()
    if not credentials['client_key'] or not credentials['client_secret'] or not credentials['account_name'] or not credentials['account_id']:
        print("Error: API credentials not found in environment variables")
        return

    # Construct the API URL using account name and ID
    api_url = f"https://{credentials['account_name']}.udemy.com/api-2.0/organizations/{credentials['account_id']}/analytics/user-course-activity/"

    # Fetch data from API
    data = fetch_api_data(api_url, credentials)
    if not data:
        return

    # Replace with your Excel file path
    excel_file = "template.xlsx"

    # Populate Excel file
    populated_file = populate_excel(data, excel_file)

    # Get SFTP configuration
    sftp_config = get_sftp_config()

    # Upload file to SFTP
    upload_to_sftp(populated_file, sftp_config)

if __name__ == "__main__":
    main()

